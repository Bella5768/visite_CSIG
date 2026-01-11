from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.conf import settings
from .models import Visiteur
import openpyxl
from openpyxl import Workbook
from io import BytesIO


@login_required
def index(request):
    search = request.GET.get('search', '')
    visiteurs = Visiteur.objects.annotate(nb_visites=Count('visites'))
    if search:
        visiteurs = visiteurs.filter(Q(nom__icontains=search) | Q(prenoms__icontains=search) | Q(telephone__icontains=search))
    paginator = Paginator(visiteurs, 20)
    return render(request, 'visiteurs/index.html', {'page_title': 'Visiteurs', 'visiteurs': paginator.get_page(request.GET.get('page', 1)), 'search_term': search})


@login_required
def ajouter(request):
    if request.method == 'POST':
        visiteur = Visiteur.objects.create(
            nom=request.POST.get('nom'), prenoms=request.POST.get('prenoms'),
            type_identite=request.POST.get('type_identite', ''),
            numero_identite=request.POST.get('numero_identite') or None,
            telephone=request.POST.get('telephone', ''), email=request.POST.get('email', ''),
            adresse=request.POST.get('adresse', '')
        )
        messages.success(request, 'Visiteur créé')
        if 'save_and_visit' in request.POST:
            return redirect('visites:nouvelle_visite_visiteur', visiteur_id=visiteur.id)
        return redirect('visiteurs:index')
    return render(request, 'visiteurs/ajouter.html', {'page_title': 'Nouveau visiteur', 'types_identite': settings.TYPES_IDENTITE})


@login_required
def modifier(request, pk):
    visiteur = get_object_or_404(Visiteur, pk=pk)
    if request.method == 'POST':
        visiteur.nom, visiteur.prenoms = request.POST.get('nom'), request.POST.get('prenoms')
        visiteur.type_identite = request.POST.get('type_identite', '')
        visiteur.numero_identite = request.POST.get('numero_identite') or None
        visiteur.telephone, visiteur.email = request.POST.get('telephone', ''), request.POST.get('email', '')
        visiteur.adresse = request.POST.get('adresse', '')
        visiteur.save()
        messages.success(request, 'Visiteur modifié')
        return redirect('visiteurs:index')
    return render(request, 'visiteurs/modifier.html', {'page_title': 'Modifier visiteur', 'visiteur': visiteur, 'types_identite': settings.TYPES_IDENTITE})


@login_required
def historique(request, pk):
    visiteur = get_object_or_404(Visiteur, pk=pk)
    return render(request, 'visiteurs/historique.html', {'page_title': f'Historique - {visiteur}', 'visiteur': visiteur, 'visites': visiteur.visites.select_related('motif', 'correspondant')})


@login_required
def supprimer(request, pk):
    visiteur = get_object_or_404(Visiteur, pk=pk)
    if request.method == 'POST':
        nom_complet = str(visiteur)
        visiteur.delete()
        messages.success(request, f'Visiteur "{nom_complet}" supprimé avec succès')
        return redirect('visiteurs:index')
    return render(request, 'visiteurs/supprimer.html', {'page_title': 'Supprimer visiteur', 'visiteur': visiteur})


@login_required
def rechercher(request):
    q = request.GET.get('q', '')
    visiteurs = Visiteur.objects.filter(Q(nom__icontains=q) | Q(prenoms__icontains=q) | Q(telephone__icontains=q)).annotate(nb_visites=Count('visites'))[:20] if q else []
    return render(request, 'visiteurs/rechercher.html', {'page_title': 'Rechercher', 'visiteurs': visiteurs, 'search_term': q})


@login_required
def api_search(request):
    q = request.GET.get('q', '')
    if len(q) < 2:
        return JsonResponse({'results': []})
    visiteurs = Visiteur.objects.filter(Q(nom__icontains=q) | Q(prenoms__icontains=q) | Q(telephone__icontains=q))[:10]
    return JsonResponse({'results': [{'id': v.id, 'nom': v.nom, 'prenoms': v.prenoms, 'telephone': v.telephone} for v in visiteurs]})


def normalize_header(h):
    """Normaliser un en-tête pour le mapping"""
    if not h:
        return ''
    import unicodedata
    h = str(h).lower().strip()
    h = unicodedata.normalize('NFD', h)
    h = ''.join(c for c in h if unicodedata.category(c) != 'Mn')
    return h


@login_required
def importer_excel(request):
    """Importer des visiteurs depuis un fichier Excel"""
    if request.method == 'POST' and request.FILES.get('fichier_excel'):
        fichier = request.FILES['fichier_excel']
        
        # Vérifier l'extension
        if not fichier.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Format de fichier non supporté. Utilisez un fichier .xlsx')
            return redirect('visiteurs:importer_excel')
        
        try:
            wb = openpyxl.load_workbook(fichier)
            ws = wb.active
            
            # Récupérer les en-têtes (première ligne)
            headers = [normalize_header(cell.value) for cell in ws[1]]
            
            # Mapper les colonnes - recherche plus flexible
            col_map = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                # Nom (pas prénom)
                if h == 'nom' or (h.startswith('nom') and 'prenom' not in h):
                    col_map['nom'] = i
                # Prénoms
                elif 'prenom' in h:
                    col_map['prenoms'] = i
                # Téléphone
                elif 'tel' in h or 'phone' in h:
                    col_map['telephone'] = i
                # Email
                elif 'mail' in h or 'email' in h:
                    col_map['email'] = i
                # Adresse
                elif 'adresse' in h:
                    col_map['adresse'] = i
                # Type identité
                elif 'type' in h:
                    col_map['type_identite'] = i
                # Numéro identité
                elif 'numero' in h or 'n°' in h or 'id' in h:
                    col_map['numero_identite'] = i
            
            # Si colonnes par position (A=Nom, B=Prénoms)
            if 'nom' not in col_map and len(headers) >= 2:
                col_map['nom'] = 0
                col_map['prenoms'] = 1
                if len(headers) >= 3:
                    col_map['telephone'] = 2
                if len(headers) >= 4:
                    col_map['email'] = 3
            
            if 'nom' not in col_map or 'prenoms' not in col_map:
                messages.error(request, f'Colonnes trouvées: {headers}. Le fichier doit contenir "Nom" et "Prénoms"')
                return redirect('visiteurs:importer_excel')
            
            # Importer les données
            created = 0
            updated = 0
            errors = 0
            error_details = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    nom = row[col_map['nom']].value
                    prenoms = row[col_map['prenoms']].value
                    
                    if not nom or not prenoms:
                        continue
                    
                    nom = str(nom).strip().upper()
                    prenoms = str(prenoms).strip().title()
                    
                    # Données optionnelles
                    telephone = ''
                    if 'telephone' in col_map:
                        val = row[col_map['telephone']].value
                        telephone = str(val).strip() if val else ''
                    
                    email = ''
                    if 'email' in col_map:
                        val = row[col_map['email']].value
                        email = str(val).strip() if val else ''
                    
                    adresse = ''
                    if 'adresse' in col_map:
                        val = row[col_map['adresse']].value
                        adresse = str(val).strip() if val else ''
                    
                    type_identite = ''
                    if 'type_identite' in col_map:
                        val = row[col_map['type_identite']].value
                        type_identite = str(val).strip() if val else ''
                    
                    numero_identite = ''
                    if 'numero_identite' in col_map:
                        val = row[col_map['numero_identite']].value
                        numero_identite = str(val).strip() if val else ''
                    
                    # Vérifier si le visiteur existe déjà
                    visiteur = None
                    if numero_identite:
                        visiteur = Visiteur.objects.filter(numero_identite=numero_identite).first()
                    
                    if not visiteur:
                        visiteur = Visiteur.objects.filter(nom__iexact=nom, prenoms__iexact=prenoms).first()
                    
                    if visiteur:
                        # Mettre à jour
                        if telephone:
                            visiteur.telephone = telephone
                        if email:
                            visiteur.email = email
                        if adresse:
                            visiteur.adresse = adresse
                        visiteur.save()
                        updated += 1
                    else:
                        # Créer
                        Visiteur.objects.create(
                            nom=nom,
                            prenoms=prenoms,
                            telephone=telephone,
                            email=email,
                            adresse=adresse,
                            type_identite=type_identite,
                            numero_identite=numero_identite if numero_identite else None
                        )
                        created += 1
                        
                except Exception as e:
                    errors += 1
                    error_details.append(f'Ligne {row_num}: {str(e)}')
                    continue
            
            if created or updated:
                messages.success(request, f'Import terminé : {created} créé(s), {updated} mis à jour, {errors} erreur(s)')
            else:
                msg = 'Aucun visiteur importé.'
                if error_details:
                    msg += f' Erreurs: {"; ".join(error_details[:3])}'
                messages.warning(request, msg)
                
        except Exception as e:
            messages.error(request, f'Erreur lors de la lecture du fichier: {str(e)}')
        
        return redirect('visiteurs:importer_excel')
    
    return render(request, 'visiteurs/importer_excel.html', {
        'page_title': 'Importer des visiteurs',
        'types_identite': settings.TYPES_IDENTITE,
    })


@login_required
def telecharger_modele_excel(request):
    """Télécharger un modèle Excel pour l'importation"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Visiteurs"
    
    # En-têtes
    headers = ['Nom', 'Prénoms', 'Téléphone', 'Email', 'Adresse', 'Type Identité', 'Numéro Identité']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="104480", end_color="104480", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Exemples
    examples = [
        ['DIALLO', 'Mamadou', '620123456', 'mamadou@email.com', 'Conakry', 'Passeport', 'P123456'],
        ['BARRY', 'Fatoumata', '621987654', 'fatoumata@email.com', 'Kindia', 'CNI', 'CNI789012'],
    ]
    for row_num, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    
    # Préparer la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=modele_import_visiteurs.xlsx'
    
    wb.save(response)
    return response


@login_required
def exporter_excel(request):
    """Exporter tous les visiteurs en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Visiteurs"
    
    # En-têtes
    headers = ['ID', 'Nom', 'Prénoms', 'Téléphone', 'Email', 'Adresse', 'Type Identité', 'Numéro Identité', 'Date création']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="104480", end_color="104480", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Données
    visiteurs = Visiteur.objects.all().order_by('nom', 'prenoms')
    for row_num, v in enumerate(visiteurs, 2):
        ws.cell(row=row_num, column=1, value=v.id)
        ws.cell(row=row_num, column=2, value=v.nom)
        ws.cell(row=row_num, column=3, value=v.prenoms)
        ws.cell(row=row_num, column=4, value=v.telephone)
        ws.cell(row=row_num, column=5, value=v.email)
        ws.cell(row=row_num, column=6, value=v.adresse)
        ws.cell(row=row_num, column=7, value=v.type_identite)
        ws.cell(row=row_num, column=8, value=v.numero_identite or '')
        ws.cell(row=row_num, column=9, value=v.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=visiteurs_export.xlsx'
    
    wb.save(response)
    return response
